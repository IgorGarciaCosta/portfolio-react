from openpyxl import load_workbook
import os

path = os.path.join(os.path.expanduser("~"), "Downloads", "OpportunitysAssessment.xlsx")
wb = load_workbook(path)
ws = wb.active
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    print(repr(row))
