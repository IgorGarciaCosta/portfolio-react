import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "Assessment"

thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
header_font = Font(bold=True, size=12)
header_fill = PatternFill("solid", fgColor="F2F2F2")

# Headers
ws["A1"] = ""
ws["B1"] = "Job Role"
ws["C1"] = "My Rating (1-5)/Motivation"
for col in ("A", "B", "C"):
    c = ws[f"{col}1"]
    c.font = header_font
    c.alignment = center
    c.fill = header_fill
    c.border = border

# Row 1 example
ws["A2"] = 1
ws["B2"] = "SE_Data Architect_Principal - Volvo Group (AB Volvo)"
ws["C2"] = ("X\n\nI have X years cognate experience in the tech stack - "
            "Java, SpringBoot....")

# Total data rows (index 1..NUM_ROWS)
NUM_ROWS = 41
# Empty rows 2..NUM_ROWS get their index number
for i in range(2, NUM_ROWS + 1):
    ws[f"A{i + 1}"] = i

for row in range(2, NUM_ROWS + 2):
    for col in ("A", "B", "C"):
        cell = ws[f"{col}{row}"]
        cell.border = border
        if col == "A":
            cell.alignment = Alignment(horizontal="left", vertical="top")
        elif col == "B":
            cell.alignment = left
        else:
            cell.alignment = center

# Column widths
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 55
ws.column_dimensions["C"].width = 25

# Row heights
ws.row_dimensions[1].height = 40
ws.row_dimensions[2].height = 110
for r in range(3, NUM_ROWS + 2):
    ws.row_dimensions[r].height = 60

downloads = os.path.join(os.path.expanduser("~"), "Downloads")
out = os.path.join(downloads, "OpportunitysAssessment.xlsx")
wb.save(out)
print(f"Saved: {out}")
