import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# (index, Job Role, Rating, Motivation)
DATA = [
    (1, "Unreal Developer / Technical Designer - RFQ9136011776", "5",
     "Over 5 years developing in Unreal Engine across three different teams. "
     "Strong in C++, Android development (inside and outside UE), the UE pipeline, "
     "and core workflows such as animation, particles, and shading. This role aligns "
     "closely with my core expertise."),

    (2, "CGI Artist Consultant (14276)", "4",
     "4+ years in 3D modeling and sculpting with Blender, Autodesk VRED, Maya, and "
     "Substance Painter. Experience developing VR applications in Unreal and Unity, "
     "plus a background in the automotive sector."),

    (3, "Senior Unreal Visualization Artist - RFQ9646012252", "4",
     "Experienced with Blender, Unreal Engine, C++/Python, and VR/AR development, and "
     "I adapt quickly to new tools.\n"
     "- Strong portfolio based on my own Unreal Engine work: YES\n"
     "- Experience with 3ds Max, Substance Suite: NO (open to learn)\n"
     "- Strong knowledge of lighting and camera work: YES\n"
     "- Receptive to learning new tools and sharing knowledge: YES\n"
     "- Understanding of materials and texturing: YES\n"
     "- CAD optimization and scene performance expertise: NO (open to learn)"),

    (4, "SE_Embedded Software Application Engineer_Expert - VOLGJPXXX38738", "5",
     "Strong C++ development in Unreal Engine and hands-on experience with Blender for "
     "modeling, rigging, and shading. My work at Ford on Unreal/C++ simulations closely "
     "mirrors this role, and I'm used to Agile/Scrum, so I would adapt very quickly."),

    (5, "Senior Generative AI Expert - RFQ6100015611", "4",
     "I work with the listed stack and with agentic AI, though not yet with full AI "
     "system development and deployment. It's an area I'm highly motivated to grow into "
     "with some dedicated study time."),

    (6, "SE_System Design Engineer_Expert - VOLGJPXXX32869", "3",
     "I'm still building my experience in systems design, so I may not meet every "
     "expert-level requirement, but I'd welcome the chance to gain more hands-on "
     "experience in this area."),

    (7, "Data Analyst - RFQ 335209938", "4",
     "A strong opportunity to grow in data analysis. With my automotive background and "
     "solid software development fundamentals, I'd welcome this as a challenge to "
     "develop further."),

    (8, "Simulation Developer - SCAN0733", "5",
     "Experience with vehicle simulations and most of the required skills. Simulation "
     "work is a field that deeply interests me, making this a great match."),

    (9, "Software Developer (Data Scientist) - SCAN0752", "5",
     "Motivated to expand within the automotive sector, which particularly interests me. "
     "Growing my data-management expertise would be a very welcome challenge."),

    (10, "SE_Data Analyst_Senior - VOLGJP00031667", "3",
     "Strong interest in Volvo, though in data analysis I feel I'd need some adaptation "
     "for this senior role."),

    (11, "CAE Engineer CFD Level 4 - VOLVJPXXX25775", "4",
     "Experience with much of the required stack, and it resembles work I've done "
     "alongside engineers at Ford. I'm only unsure whether my Computer Engineering "
     "degree (rather than Electrical) fully applies."),

    (12, "SE_Simulation & Analysis Engineer B_Level 4 - VOLGJP00032332", "5",
     "Simulation work motivates me. This role differs from my current work, and "
     "combining technologies I already know with entirely new ones is highly motivating."),

    (13, "Senior Thermal Engineer - LP 221", "2",
     "I don't have expertise in the required software or the listed responsibilities, "
     "so my profile likely isn't the right fit for this role."),

    (14, "CAE Engineer KROCK Level 4 - VOLVJP00025766", "3",
     "The concept interests me, but I currently don't match much of the stack. Even so, "
     "it's a challenge I would gladly take on."),

    (15, "Software Engineer Level 3 (Frontend) - VOLVXXXXX25326", "5",
     "One of my top choices. My stack is aligned and I'm currently taking an additional "
     "AWS course. A great chance to fully dive into front-end development."),

    (16, "SE_Data Analyst_Associate - VOLGXXXXX29820", "3",
     "Data analysis would be a good area for me to grow into."),

    (17, "Software/Data Engineer (GCP)", "3",
     "I'm not as aligned with the requirements and stack here as I am with other roles."),

    (18, "Software Engineer (C# and .NET)", "4",
     "I cover the skills and stack well. A great opportunity to broaden my backend work, "
     "which I actively pursue through study and personal projects."),

    (19, "Senior AI Platform / Data Engineer", "2",
     "Some interest, but as a senior role in an area I'd need study time for, it isn't "
     "the best match for what I'm looking for right now."),

    (20, "Fullstack Developer (Next.js / NestJS)", "4",
     "Strong interest, my stack is aligned, and there's plenty of room to grow."),

    (21, "Software Engineer Level 2 [EE] - VOLVJPXXX25855", "5",
     "I know the stack relatively well and I'm very motivated to learn the remaining "
     "required technologies."),

    (22, "SE_Embedded Software Application Engineer_Experienced - VOLGJP00040016", "3",
     "Experience in test development for applications and in the automotive sector. "
     "A good area to apply what I know while learning new things."),

    (23, "Senior Data Scientist - Python & MATLAB", "3",
     "I'm interested in AI development, but I'm not sure I'd be a strong fit here given "
     "my current level of expertise in this specific area."),

    (24, "Software Developer - Car Cloud Platform - JX-50032", "4",
     "Strong interest in cloud in this context. It's a fast-growing area I'd very much "
     "like to work in."),

    (25, "Product Expert AI", "1",
     "This doesn't align closely with my area of work."),

    (26, "Fullstack Developer - Python", "4",
     "Experience with Python, including backend and personal projects, plus solid "
     "experience with front-end libraries such as Node.js and React."),

    (27, "Senior Agentic AI Developer", "3",
     "Experience using agents, though not in building AI itself. Developing an AI from "
     "scratch would be a new field and a welcome challenge."),

    (28, "Senior Project Manager - Information Security & AI (16329)", "2",
     "I'm not at a project-manager stage professionally, at least not in an area where "
     "I still have more to develop."),

    (29, "SE_System Design Engineer_Expert - VOLGJPXXX42783", "1",
     "This role focuses far more on automotive/thermal engineering than on software "
     "development, so it doesn't fit my area well."),

    (30, "SE_System Design Engineer_Senior Expert (.NET/C#) - VOLGJPXXX42606", "3",
     "Interested in both the role and the .NET/C# stack, which aligns with my backend "
     "experience."),

    (31, "AI Tool Specialist", "3",
     "Interested, with solid experience using enterprise AI tools during my time at Ford."),

    (32, "Surface Designer - SCAN01303", "5",
     "Hard-surface 3D modeling is one of my strongest interests. I don't yet have "
     "experience with Alias (required), but I'm very motivated and quick to learn it."),

    (33, "Lead Digital Modeler - RFQ 9645011925", "5",
     "3D modeling is a core interest of mine. Alias isn't in my toolkit yet, but I'm "
     "eager and quick to pick up new modeling tools."),

    (34, "Sr. Component Designer - RFQ9730011643-45", "5",
     "Strong passion for 3D and product design. I'd need to learn Alias, which the role "
     "requires, but I'm confident I'd ramp up fast."),

    (35, "Jr Designer CMF Program - RFQ 9730012054", "4",
     "Slightly lower rating as it's a junior position, but it follows the same principle "
     "as the other 3D roles that interest me."),

    (36, "Senior UI/UX Artist with Unity (Mobile Games)", "5",
     "I've worked professionally with digital design, community, and UI creation, so this "
     "is a role where I believe I'd perform very well."),

    (37, "Fullstack .NET Developer with Backend Focus", "3",
     "Interested in fullstack roles, and this one includes Unity testing and Scrum, where "
     "I have experience. I did note it requires Swedish and more total experience than I "
     "currently have."),

    (38, "Software .NET Fullstack Developer - JX-47267", "4",
     "Works with technologies I'm currently studying and want to build a stronger "
     "foundation in, making it a perfect opportunity. It also lists AI integration, Unity, "
     "and Unreal as nice-to-haves, all of which I have experience with."),

    (39, "Visual Designer (Marketing Artist)", "3",
     "I enjoy working with engines, but my contribution here would lean more toward "
     "development than the design focus this role requires."),

    (40, "CGI Visualization Artist - SCAN01415", "5",
     "Strong interest in vehicle visualization, especially given the stack required for "
     "this role (VRED, Blender, Unreal, Unity)."),

    (41, "Embedded Developer - Linköping", "3",
     "Interested in growing my C++ work, though I haven't done embedded projects for some "
     "time and would need a short ramp-up period."),

    (42, "SE_Geometrical Architect_Senior - VOLGJP00041951", "2",
     "This doesn't seem to fit my field of work very well."),
]

wb = Workbook()
ws = wb.active
ws.title = "Assessment"

thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="top", wrap_text=True)
idx_align = Alignment(horizontal="center", vertical="top")
header_font = Font(bold=True, size=12)
header_fill = PatternFill("solid", fgColor="F2F2F2")

# Headers
ws["A1"] = "#"
ws["B1"] = "Job Role"
ws["C1"] = "My Rating (1-5) / Motivation"
for col in ("A", "B", "C"):
    c = ws[f"{col}1"]
    c.font = header_font
    c.alignment = center
    c.fill = header_fill
    c.border = border

B_WIDTH = 42
C_WIDTH = 68


def est_lines(text, width):
    lines = 0
    for para in text.split("\n"):
        lines += max(1, -(-len(para) // width))  # ceil division
    return lines


for i, (idx, role, rating, motivation) in enumerate(DATA):
    row = i + 2
    ws[f"A{row}"] = idx
    ws[f"B{row}"] = role
    ws[f"C{row}"] = f"{rating}\n\n{motivation}"

    ws[f"A{row}"].alignment = idx_align
    ws[f"B{row}"].alignment = left
    ws[f"C{row}"].alignment = left
    for col in ("A", "B", "C"):
        ws[f"{col}{row}"].border = border

    lines = max(
        est_lines(role, B_WIDTH),
        est_lines(f"{rating}\n\n{motivation}", C_WIDTH),
    )
    ws.row_dimensions[row].height = max(30, lines * 15 + 6)

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = B_WIDTH
ws.column_dimensions["C"].width = C_WIDTH
ws.row_dimensions[1].height = 30

downloads = os.path.join(os.path.expanduser("~"), "Downloads")
out = os.path.join(downloads, "OpportunitysAssessment2.xlsx")
wb.save(out)
print(f"Saved: {out}  ({len(DATA)} rows)")
